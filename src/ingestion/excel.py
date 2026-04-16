"""Excel ingestor using openpyxl.

Converts each worksheet into a page image rendered as a grid and extracts
cell data as a structured TableData representation. Handles merged cells.
"""
from __future__ import annotations

import logging
from typing import Any

import numpy as np
import openpyxl
from openpyxl.utils.cell import get_column_letter
from PIL import Image, ImageDraw, ImageFont

from src.ingestion.base import IngestorBase, PageResult

logger = logging.getLogger(__name__)

_PAGE_WIDTH_PX: int = 2480
_PAGE_HEIGHT_PX: int = 3508
_MARGIN_PX: int = 80
_CELL_HEIGHT_PX: int = 40
_CELL_WIDTH_PX: int = 200
_FONT_SIZE: int = 20
_DPI: int = 300


def _load_font() -> Any:
    """Load a TrueType font or fall back to PIL default.

    Returns:
        ImageFont: font object for rendering.
    """
    try:
        return ImageFont.truetype(
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", _FONT_SIZE
        )
    except (IOError, OSError):
        return ImageFont.load_default()


def _cell_str(value: Any) -> str:
    """Convert a cell value to a stripped string, returning empty for None.

    Returns:
        str: string representation of the cell value.
    """
    if value is None:
        return ""
    return str(value).strip()


def _build_table_data_dict(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict:
    """Extract worksheet data into a TableData-compatible dictionary.

    The first non-empty row is treated as the header row. Merged cell
    regions are resolved so each logical cell carries the top-left value
    of its merge group.

    Returns:
        dict: dict with keys 'headers', 'rows', 'cells', 'markdown'.
    """
    merged_lookup: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        top_left_cell = ws.cell(merged_range.min_row, merged_range.min_col)
        for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
            for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                merged_lookup[(row_idx, col_idx)] = _cell_str(top_left_cell.value)

    all_rows: list[list[str]] = []
    for row in ws.iter_rows():
        row_values: list[str] = []
        for cell in row:
            if (cell.row, cell.column) in merged_lookup:
                row_values.append(merged_lookup[(cell.row, cell.column)])
            else:
                row_values.append(_cell_str(cell.value))
        if any(v for v in row_values):
            all_rows.append(row_values)

    if not all_rows:
        return {"headers": [], "rows": [], "cells": [], "markdown": ""}

    headers = all_rows[0]
    data_rows = all_rows[1:]

    cells: list[dict] = []
    for col_idx, header in enumerate(headers):
        cells.append(
            {
                "row": 0,
                "col": col_idx,
                "row_span": 1,
                "col_span": 1,
                "value": header,
                "is_header": True,
            }
        )
    for row_idx, row in enumerate(data_rows, start=1):
        for col_idx, value in enumerate(row):
            cells.append(
                {
                    "row": row_idx,
                    "col": col_idx,
                    "row_span": 1,
                    "col_span": 1,
                    "value": value,
                    "is_header": False,
                }
            )

    md_lines: list[str] = []
    md_lines.append("| " + " | ".join(headers) + " |")
    md_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for row in data_rows:
        md_lines.append("| " + " | ".join(row) + " |")
    markdown = "\n".join(md_lines)

    return {
        "headers": headers,
        "rows": data_rows,
        "cells": cells,
        "markdown": markdown,
    }


def _render_sheet_to_image(ws: openpyxl.worksheet.worksheet.Worksheet) -> np.ndarray:
    """Render a worksheet as a grid image on a white A4 canvas.

    Only as many rows and columns as fit within the page boundaries are
    drawn. Cell values are truncated if they exceed the column width.

    Returns:
        np.ndarray: BGR uint8 image.
    """
    pil_image = Image.new("RGB", (_PAGE_WIDTH_PX, _PAGE_HEIGHT_PX), color=(255, 255, 255))
    draw = ImageDraw.Draw(pil_image)
    font = _load_font()

    usable_width = _PAGE_WIDTH_PX - 2 * _MARGIN_PX
    usable_height = _PAGE_HEIGHT_PX - 2 * _MARGIN_PX

    max_cols = usable_width // _CELL_WIDTH_PX
    max_rows = usable_height // _CELL_HEIGHT_PX

    for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows)):
        if row_idx >= max_rows:
            break
        y_top = _MARGIN_PX + row_idx * _CELL_HEIGHT_PX
        for col_idx, cell in enumerate(row):
            if col_idx >= max_cols:
                break
            x_left = _MARGIN_PX + col_idx * _CELL_WIDTH_PX
            draw.rectangle(
                [x_left, y_top, x_left + _CELL_WIDTH_PX, y_top + _CELL_HEIGHT_PX],
                outline=(180, 180, 180),
            )
            cell_text = _cell_str(cell.value)
            draw.text(
                (x_left + 4, y_top + 4),
                cell_text[:24],
                fill=(0, 0, 0),
                font=font,
            )

    return np.array(pil_image)[:, :, ::-1].copy()


class ExcelIngestor(IngestorBase):
    """Ingestor for Microsoft Excel workbooks (.xlsx, .xls, .xlsm).

    Each worksheet is rendered as a grid page image. The full cell data
    is stored in metadata['table_data'] as a dictionary compatible with
    the TableData model. All cell values are concatenated as native_text.
    """

    @property
    def supported_extensions(self) -> list[str]:
        """File extensions handled by this ingestor.

        Returns:
            list[str]: ['.xlsx', '.xls', '.xlsm']
        """
        return [".xlsx", ".xls", ".xlsm"]

    def ingest(self, source_path: str) -> list[PageResult]:
        """Load a workbook and return one PageResult per worksheet.

        Returns:
            list[PageResult]: ordered list of page results, 1-indexed page_number.
        """
        wb = openpyxl.load_workbook(source_path, data_only=True)
        results: list[PageResult] = []

        for page_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            table_data_dict = _build_table_data_dict(ws)
            image = _render_sheet_to_image(ws)

            all_values: list[str] = []
            for row in ws.iter_rows():
                for cell in row:
                    v = _cell_str(cell.value)
                    if v:
                        all_values.append(v)
            native_text = "\n".join(all_values)

            results.append(
                PageResult(
                    page_number=page_idx + 1,
                    image=image,
                    native_text=native_text,
                    estimated_dpi=_DPI,
                    quality_tier="standard",
                    is_scanned=False,
                    metadata={
                        "sheet_name": sheet_name,
                        "table_data": table_data_dict,
                    },
                )
            )

        return results
